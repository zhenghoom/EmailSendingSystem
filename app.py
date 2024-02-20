import time

from flask import Flask, render_template, request, url_for, redirect
from time import sleep
import win32com.client as client
import pythoncom
import openpyxl
from tkinter import *
from tkinter import messagebox
import csv

app = Flask(__name__, static_url_path="/static")

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/mass_send', methods=['POST'])
def saveValue():
    file = request.form['file']
    content = request.form['content']
    chunks = request.form['num-chunks']
    time = request.form['time-interval']

    pythoncom.CoInitialize()
    with open(file, newline='') as lines:
        reader = csv.reader(lines)
        individual = [row for row in reader]

    # Emails to send out every cycle
    chunks = [individual[x:x + int(chunks)] for x in range(0, len(individual), int(chunks))]

    template = content
    outlook = client.Dispatch("Outlook.Application")
    if request.form['submitType'] == 'Submit':
        for chunk in chunks:
            for name, email, subject, attachment in chunk:
                message = outlook.CreateItem(0)
                # message.Display()
                message.To = email
                message.Subject = subject
                #Get signature
                # message.GetInspector
                index = message.HTMLbody.find('>', message.HTMLbody.find('<body'))
                message.HTMLBody = message.HTMLbody[:index + 1] + template.format(name)

                attachments = attachment.split(', ')
                for i in range(len(attachments)):
                    message.Attachments.Add(attachments[i])
                message.Send()
            sleep(int(time))
        return render_template('loading.html'), {"Refresh": "4; url=mass_send"}

    elif request.form['submitType'] == 'Save':
        for chunk in chunks:
            for name, email, subject, attachment in chunk:
                message = outlook.CreateItem(0)
                message.To = email
                message.Subject = subject
                # Get signature
                message.GetInspector
                index = message.HTMLbody.find('>', message.HTMLbody.find('<body'))
                message.HTMLBody = message.HTMLbody[:index + 1] + template.format(name) + message.HTMLbody[index + 1:]
                attachments = attachment.split(', ')
                for i in range(len(attachments)):
                    message.Attachments.Add(attachments[i])
                message.Save()
            sleep(int(time))
    #     if attachment == '':
    #         for chunk in chunks:
    #             for name, email in chunk:
    #                 message = outlook.CreateItem(0)
    #                 message.To = email
    #                 message.Subject = subject
    #                 message.HTMLBody = template.format(name)
    #                 message.Save()
    #             # sleep(int(time))
    #     else:
    #         for chunk in chunks:
    #             for name, email in chunk:
    #                 message = outlook.CreateItem(0)
    #                 message.To = email
    #                 message.Subject = subject
    #                 message.HTMLBody = template.format(name)
    #                 for i in range(len(attachments)):
    #                     message.Attachments.Add(attachments[i])
    #                 message.Save()
    #             # sleep(int(time))
        return render_template('loading.html'), {"Refresh": "4; url=mass_send"}

@app.route('/mass_send')
def massSend():
    return render_template('massSend.html')

@app.route('/excel')
def excelTemp():
    return render_template('excelTemp.html')

@app.route('/excel', methods=['POST'])
def excelFunc():
    excelfile = request.form['excelfile']
    reference = request.form['reference']
    main = request.form['main']
    vendor = request.form['vendor']
    matno = request.form['matno']
    quolink1 = request.form['quolink1']
    quolink2 = request.form['quolink2']
    plant = request.form['plant']
    venmatno = request.form['venmatno']
    orderunit = request.form['orderunit']
    leadtime = request.form['leadtime']
    purchaseGroup = request.form['purchaseGroup']
    moq = request.form['moq']
    netprice = request.form['netprice']

    excelFile = openpyxl.load_workbook(excelfile)
    reference_sheet = excelFile[reference]
    main_sheet =excelFile[main]
    for j in reference_sheet.iter_rows():
        if int(vendor) == j[1].value:
            main_sheet.cell(column=5, row=main_sheet.max_row+1, value=vendor)
            main_sheet.cell(row=main_sheet.max_row, column=4).value = j[0].value  # vendor code
            main_sheet.cell(row=main_sheet.max_row, column=20).value = j[2].value  # incoterm
            main_sheet.cell(row=main_sheet.max_row, column=21).value = j[3].value   #incoterm desc
            main_sheet.cell(row=main_sheet.max_row, column=17).value = j[4].value  # currency
            main_sheet.cell(row=main_sheet.max_row, column=6).value = matno
            main_sheet.cell(row=main_sheet.max_row, column=8).value = quolink1
            main_sheet.cell(row=main_sheet.max_row, column=9).value = quolink2
            main_sheet.cell(row=main_sheet.max_row, column=10).value = plant
            main_sheet.cell(row=main_sheet.max_row, column=11).value = venmatno
            main_sheet.cell(row=main_sheet.max_row, column=12).value = orderunit
            main_sheet.cell(row=main_sheet.max_row, column=13).value = leadtime
            main_sheet.cell(row=main_sheet.max_row, column=14).value = purchaseGroup
            main_sheet.cell(row=main_sheet.max_row, column=15).value = moq
            main_sheet.cell(row=main_sheet.max_row, column=16).value = netprice

    # main_sheet.delete_rows(9,1)

            tdytime = time.strftime("%Y%m%d")
            txtfile = open(tdytime + ".txt", 'a')
            txtfile.write("\t" + "\t" + "\t" + str(j[0].value) + "\t" + str(j[1].value) + "\t" + matno + "\t" + "\t" + quolink1 + "\t" + quolink2 + "\t" + plant + "\t" + venmatno
                          + "\t" + orderunit + "\t" + leadtime + "\t" + purchaseGroup + "\t" + moq + "\t" + netprice + "\t" + str(j[4].value) + "\t" + "\t" + "\t" + str(j[2].value) +
                          "\t" + str(j[3].value) + "\n")
            txtfile.close()
            excelFile.save(excelfile)

        # elif int(vendor) != j[1].value:
        #     return messagebox.showerror("Vendor code", "Vendor code not found")



    return render_template('loading.html'), {"Refresh": "4; url=excel"}
@app.route('/loading')
def submitLoading():
    return render_template('submitLoading.html'),{"Refresh": "1; url=/"}

if __name__ == '__main__':
    app.run(debug=True)
